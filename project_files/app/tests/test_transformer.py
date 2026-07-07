from __future__ import annotations

import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from sap_xlsx_transformer.rules import load_rules
from sap_xlsx_transformer.transformer import transform_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[3]
SOURCE_FILE = PROJECT_ROOT / "Источники" / "inbox" / "100019128.XLSX"
CLEAN_SOURCE_FILE = PROJECT_ROOT / "Источники" / "inbox" / "тестовая табличка_0.xlsx"


class TransformerTest(unittest.TestCase):
    def test_transforms_example_file(self) -> None:
        rules = load_rules()
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_file = Path(tmp_dir) / "output.xlsx"

            summary = transform_workbook(
                SOURCE_FILE,
                output_file,
                rules,
                calc_date=date(2026, 7, 1),
            )

            self.assertEqual(summary.processed_dr_rows, 1013)
            self.assertEqual(summary.base_formula_rows, 1013)
            self.assertTrue(output_file.exists())

            workbook = load_workbook(output_file, data_only=False)
            worksheet = workbook.active

            self.assertEqual(worksheet.max_column, 28)
            self.assertEqual(worksheet.max_row, 1775)
            self.assertIsNone(worksheet.auto_filter.ref)
            self.assertFalse(any(worksheet.row_dimensions[row].hidden for row in range(1, 1776)))

            self.assertEqual(worksheet["T339"].value, "=MAX(0,P339-O339)")
            self.assertEqual(worksheet["U339"].value, 0.001)
            self.assertEqual(worksheet["V339"].value, "=U339*T339*R339")
            self.assertIsNone(worksheet["X339"].value)
            self.assertIsNone(worksheet["AA339"].value)

            self.assertEqual(worksheet["T703"].value, "=MAX(0,$AB$1-O703)")
            self.assertIsNone(worksheet["X703"].value)
            self.assertIsNone(worksheet["T2"].value)

    def test_transforms_clean_sap_input_without_existing_calculation_properties(self) -> None:
        rules = load_rules()
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_file = Path(tmp_dir) / "clean-output.xlsx"

            summary = transform_workbook(
                CLEAN_SOURCE_FILE,
                output_file,
                rules,
                calc_date=date(2026, 7, 1),
            )

            self.assertEqual(summary.processed_dr_rows, 1013)
            self.assertTrue(output_file.exists())

            workbook = load_workbook(output_file, data_only=False)
            worksheet = workbook.active

            self.assertEqual(worksheet.max_column, 28)
            self.assertEqual(worksheet["T7"].value, "=MAX(0,P7-O7)")
            self.assertEqual(worksheet["V7"].value, "=U7*T7*R7")
            self.assertEqual(worksheet["W7"].value, "=V7+R7")
            self.assertIsNone(worksheet["X7"].value)
            self.assertIsNone(worksheet["T2"].value)


if __name__ == "__main__":
    unittest.main()
