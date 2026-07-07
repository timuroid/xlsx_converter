from __future__ import annotations

from io import BytesIO
from pathlib import Path
import unittest

from openpyxl import load_workbook

from sap_xlsx_transformer.web import create_app


PROJECT_ROOT = Path(__file__).resolve().parents[3]
SOURCE_FILE = PROJECT_ROOT / "Источники" / "inbox" / "100019128.XLSX"
CLEAN_SOURCE_FILE = PROJECT_ROOT / "Источники" / "inbox" / "тестовая табличка_0.xlsx"


class WebAppTest(unittest.TestCase):
    def test_index_loads(self) -> None:
        app = create_app()
        app.config["TESTING"] = True

        with app.test_client() as client:
            response = client.get("/")

        self.assertEqual(response.status_code, 200)
        self.assertIn("Калькулятор КК SAP XLSX".encode(), response.data)

    def test_convert_returns_transformed_xlsx(self) -> None:
        app = create_app()
        app.config["TESTING"] = True

        with SOURCE_FILE.open("rb") as file:
            payload = {
                "calc_date": "2026-07-01",
                "file": (BytesIO(file.read()), "100019128.xlsx"),
            }

        with app.test_client() as client:
            response = client.post("/convert", data=payload, content_type="multipart/form-data")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.mimetype,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("100019128_converted.xlsx", response.headers["Content-Disposition"])

        workbook = load_workbook(BytesIO(response.data), data_only=False)
        worksheet = workbook.active
        self.assertEqual(worksheet["T339"].value, "=MAX(0,P339-O339)")
        self.assertIsNone(worksheet["X339"].value)
        self.assertEqual(worksheet["T703"].value, "=MAX(0,$AB$1-O703)")
        self.assertIsNone(worksheet["T2"].value)

    def test_convert_accepts_clean_sap_input(self) -> None:
        app = create_app()
        app.config["TESTING"] = True

        with CLEAN_SOURCE_FILE.open("rb") as file:
            payload = {
                "calc_date": "2026-07-01",
                "file": (BytesIO(file.read()), "clean-input.xlsx"),
            }

        with app.test_client() as client:
            response = client.post("/convert", data=payload, content_type="multipart/form-data")

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.data), data_only=False)
        worksheet = workbook.active
        self.assertEqual(worksheet.max_column, 28)
        self.assertEqual(worksheet["T7"].value, "=MAX(0,P7-O7)")

    def test_convert_shows_processing_error(self) -> None:
        app = create_app()
        app.config["TESTING"] = True

        payload = {
            "calc_date": "2026-07-01",
            "file": (BytesIO(b"not an xlsx"), "broken.xlsx"),
        }

        with app.test_client() as client:
            response = client.post(
                "/convert",
                data=payload,
                content_type="multipart/form-data",
                follow_redirects=True,
            )

        self.assertEqual(response.status_code, 200)
        self.assertIn("Не удалось обработать файл:".encode(), response.data)


if __name__ == "__main__":
    unittest.main()
