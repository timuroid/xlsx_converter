from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import column_index_from_string, coordinate_to_tuple, get_column_letter
from openpyxl.workbook.properties import CalcProperties
from openpyxl.worksheet.worksheet import Worksheet

from .rules import TransformerRules


BLUE_HEADER_FILL = PatternFill(fill_type="solid", fgColor="BDD7EE")
BLUE_DATA_FILL = PatternFill(fill_type="solid", fgColor="DDEBF7")
SERVICE_HEADER_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")


@dataclass(frozen=True)
class TransformSummary:
    output_path: Path
    worksheet_name: str
    max_row: int
    processed_dr_rows: int
    base_formula_rows: int


def transform_workbook(
    input_path: str | Path,
    output_path: str | Path,
    rules: TransformerRules,
    calc_date: date | None = None,
) -> TransformSummary:
    source_path = Path(input_path)
    target_path = Path(output_path)
    target_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(source_path)
    worksheet = workbook.active

    calculation_date = calc_date or date.today()
    _prepare_workbook_for_recalculation(workbook)
    _remove_columns_after_sap_area(worksheet, rules)
    _write_headers_and_formats(worksheet, rules, calculation_date)

    max_data_row = _find_last_data_row(worksheet, rules)
    _delete_rows_after_data(worksheet, max_data_row)
    processed_dr_rows = 0
    base_formula_rows = 0

    for row in range(rules.sheet.first_data_row, max_data_row + 1):
        document_type = worksheet[f"{rules.filters.document_type_column}{row}"].value
        if document_type == rules.filters.process_document_type:
            processed_dr_rows += 1
            _write_base_formulas(worksheet, row, rules)
            base_formula_rows += 1
        elif rules.logic.clear_calculation_cells_for_non_dr:
            _clear_calculation_cells(worksheet, row, rules)

    _apply_output_layout(worksheet, rules, max_data_row)
    if not rules.logic.keep_source_filters:
        worksheet.auto_filter.ref = None
        if rules.logic.unhide_rows_when_clearing_filters:
            _unhide_rows(worksheet, max_data_row)

    workbook.save(target_path)

    return TransformSummary(
        output_path=target_path,
        worksheet_name=worksheet.title,
        max_row=max_data_row,
        processed_dr_rows=processed_dr_rows,
        base_formula_rows=base_formula_rows,
    )


def _prepare_workbook_for_recalculation(workbook: Any) -> None:
    if workbook.calculation is None:
        workbook.calculation = CalcProperties()
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True


def _remove_columns_after_sap_area(worksheet: Worksheet, rules: TransformerRules) -> None:
    sap_last_index = column_index_from_string(rules.sheet.sap_last_column)
    if worksheet.max_column > sap_last_index:
        worksheet.delete_cols(sap_last_index + 1, worksheet.max_column - sap_last_index)


def _write_headers_and_formats(
    worksheet: Worksheet,
    rules: TransformerRules,
    calculation_date: date,
) -> None:
    header_row = rules.sheet.header_row
    sap_last_index = column_index_from_string(rules.sheet.sap_last_column)
    source_header = worksheet.cell(header_row, sap_last_index)
    source_border = copy(source_header.border)
    source_alignment = copy(source_header.alignment)
    source_protection = copy(source_header.protection)

    for column, config in rules.columns.items():
        cell = worksheet[f"{column}{header_row}"]
        cell.value = config["header"]
        cell.fill = BLUE_HEADER_FILL
        cell.border = copy(source_border)
        cell.alignment = copy(source_alignment)
        cell.protection = copy(source_protection)

    calc_date_cell = worksheet[rules.sheet.calculation_date_cell]
    calc_date_cell.value = datetime.combine(calculation_date, datetime.min.time())
    calc_date_cell.number_format = "yyyy-mm-dd"
    calc_date_cell.fill = SERVICE_HEADER_FILL
    calc_date_cell.border = copy(source_border)
    calc_date_cell.alignment = copy(source_alignment)


def _find_last_data_row(worksheet: Worksheet, rules: TransformerRules) -> int:
    sap_last_index = column_index_from_string(rules.sheet.sap_last_column)
    for row in range(worksheet.max_row, rules.sheet.first_data_row - 1, -1):
        for column_index in range(1, sap_last_index + 1):
            if worksheet.cell(row, column_index).value not in (None, ""):
                return row
    return rules.sheet.first_data_row - 1


def _write_base_formulas(worksheet: Worksheet, row: int, rules: TransformerRules) -> None:
    payment_date = rules.inputs.payment_date_column
    clearing_date = rules.inputs.clearing_date_column
    amount = rules.inputs.amount_column
    calc_date = _absolute_cell_reference(rules.sheet.calculation_date_cell)

    if worksheet[f"{clearing_date}{row}"].value in (None, ""):
        worksheet[f"T{row}"] = f"=MAX(0,{calc_date}-{payment_date}{row})"
    else:
        worksheet[f"T{row}"] = f"=MAX(0,{clearing_date}{row}-{payment_date}{row})"

    worksheet[f"U{row}"] = rules.penalty_rate
    worksheet[f"V{row}"] = f"=U{row}*T{row}*{amount}{row}"
    worksheet[f"W{row}"] = f"=V{row}+{amount}{row}"


def _clear_calculation_cells(worksheet: Worksheet, row: int, rules: TransformerRules) -> None:
    for column in rules.calculation_columns:
        worksheet[f"{column}{row}"] = None


def _delete_rows_after_data(worksheet: Worksheet, max_data_row: int) -> None:
    if worksheet.max_row > max_data_row:
        worksheet.delete_rows(max_data_row + 1, worksheet.max_row - max_data_row)


def _apply_output_layout(worksheet: Worksheet, rules: TransformerRules, max_data_row: int) -> None:
    for column in rules.calculation_columns:
        worksheet.column_dimensions[column].width = _default_width_for_column(column)

    for row in range(rules.sheet.first_data_row, max_data_row + 1):
        for column in rules.calculation_columns:
            cell = worksheet[f"{column}{row}"]
            cell.fill = BLUE_DATA_FILL
            cell.alignment = Alignment(horizontal="right" if column not in ("U",) else "center")

    for column in ("T",):
        if column in rules.columns:
            for row in range(rules.sheet.first_data_row, max_data_row + 1):
                worksheet[f"{column}{row}"].number_format = "0"

    for column in ("V", "W"):
        if column in rules.columns:
            for row in range(rules.sheet.first_data_row, max_data_row + 1):
                worksheet[f"{column}{row}"].number_format = "#,##0.00"

    if "U" in rules.columns:
        for row in range(rules.sheet.first_data_row, max_data_row + 1):
            worksheet[f"U{row}"].number_format = "0.00%"

    if worksheet.freeze_panes is None:
        worksheet.freeze_panes = "A2"


def _unhide_rows(worksheet: Worksheet, max_data_row: int) -> None:
    for row in range(1, max_data_row + 1):
        worksheet.row_dimensions[row].hidden = False


def _absolute_cell_reference(cell_reference: str) -> str:
    row, column = coordinate_to_tuple(cell_reference)
    return f"${get_column_letter(column)}${row}"


def _default_width_for_column(column: str) -> float:
    widths = {
        "T": 18,
        "U": 10,
        "V": 14,
        "W": 14,
    }
    return widths.get(column, 14)
